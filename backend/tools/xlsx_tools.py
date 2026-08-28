"""Real Excel-spreadsheet generation tool, sandboxed to backend/workspace/.

Produces actual .xlsx files via openpyxl so the agent can deliver
spreadsheets alongside Word documents and presentations.

Expected spreadsheet structure (see xlsx_generate):
    Sheet title -> the `title` argument (truncated to Excel's 31-char limit)
    Row 1       -> bold header row from `headers`
    Rows 2..N   -> one row per entry in `rows`

Future extension points (not implemented yet):
    - multiple worksheets
    - cell formatting (colors, number formats)
    - formulas and charts
The row-writing loop is isolated so those additions only extend it.
"""

import re
import sys
from pathlib import Path

# Make `tools` importable both when this module is imported by the app
# (backend on sys.path) and when run directly (backend/tools on sys.path).
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from tools.file_tools import WORKSPACE_DIR, _resolve_safe  # noqa: E402

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    # openpyxl missing: xlsx_generate reports a clean error at call time
    # instead of failing the whole module import.
    Workbook = None
    Font = None


# Excel's hard limit on worksheet title length.
MAX_SHEET_TITLE = 31


def _slugify(title: str) -> str:
    """Turn a title into a safe default filename: lowercase, underscores."""
    slug = title.lower().strip()
    slug = re.sub(r"[^a-z0-9]+", "_", slug)
    slug = slug.strip("_")
    return slug or "spreadsheet"


def _flatten_headers(headers) -> list | None:
    """Normalize the model's headers output into a flat list of strings.

    The small model sometimes emits headers as a list of lists (it copies
    the "rows are lists" shape) or mixes dicts in. Flatten one nesting
    level and stringify, so the tool still produces a usable file.
    Returns None if nothing usable is left.
    """
    flat = []
    for item in headers:
        if isinstance(item, (list, tuple)):
            flat.extend(str(x) for x in item)
        elif isinstance(item, dict):
            flat.extend(f"{k}: {v}" for k, v in item.items())
        else:
            flat.append(item)
    return flat or None


def _coerce_row(row) -> list | None:
    """Coerce a row into a flat list of cell values, or None if impossible.

    Accepts a list (already correct) or a dict (keys as cells, values
    dropped - the model sometimes emits dict rows for xlsx).
    """
    if isinstance(row, list):
        return row
    if isinstance(row, dict):
        return list(row.keys())
    return None


def _validate_inputs(headers, rows) -> str | None:
    """Return an error message if headers/rows are unusable, else None.

    headers must yield a non-empty list of strings; rows must be a
    non-empty list of lists (dict rows are coerced). Length mismatches are
    NOT an error here - they are handled by padding in xlsx_generate.
    """
    if not isinstance(headers, list) or not headers:
        return "headers must be a non-empty list of column names."

    flattened = _flatten_headers(headers)
    if not flattened or any(
        not isinstance(h, str) or not str(h).strip() for h in flattened
    ):
        return "headers must contain non-empty column name strings."

    if not isinstance(rows, list) or not rows:
        return "rows must be a non-empty list of row lists."

    for index, row in enumerate(rows, 1):
        if _coerce_row(row) is None:
            return f"row {index} is not a list or dict: {row!r}"
    return None


def _normalize_rows(headers, rows) -> tuple[list, int]:
    """Pad/truncate each row to the header length; count how many changed.

    The small model sometimes emits rows with the wrong number of cells.
    Padding short rows with empty strings (and truncating long ones) keeps
    the file valid instead of failing the whole tool call.
    """
    width = len(headers)
    normalized = []
    adjusted = 0
    for row in rows:
        if len(row) != width:
            adjusted += 1
            row = row[:width] + [""] * (width - len(row))
        normalized.append(row)
    return normalized, adjusted


def xlsx_generate(
    title: str,
    headers: list,
    rows: list,
    filename: str = None,
    provenance_record: dict = None,
) -> dict:
    """Create an Excel spreadsheet in the workspace and save it.

    Args:
        title: Spreadsheet title; also used as the worksheet name,
            truncated to Excel's 31-character limit.
        headers: List of column name strings.
        rows: List of lists, each matching `headers` in order. Rows with
            the wrong length are padded/truncated to fit (a warning is
            logged when this happens).
        filename: Optional output name. Defaults to a slug of the title
            (e.g. "Inspection Log" -> "inspection_log.xlsx"). ".xlsx" is
            appended if missing.
        provenance_record: Optional provenance metadata dict. When provided,
            adds a second worksheet tab named "Provenance".

    Returns:
        {"status": "success", "output": "Spreadsheet saved as <name>",
         "file_path": "<full path>"} on success, or
        {"status": "error", "output": "<clear error message>"} on failure.
        Never raises.
    """
    if Workbook is None:
        return {
            "status": "error",
            "output": "openpyxl is not installed. Run: pip install openpyxl",
        }

    error = _validate_inputs(headers, rows)
    if error:
        return {"status": "error", "output": error}

    headers = _flatten_headers(headers)
    rows = [_coerce_row(row) for row in rows]

    if not isinstance(title, str) or not title.strip():
        return {"status": "error", "output": "Title must be a non-empty string."}

    if filename is None or not str(filename).strip():
        filename = _slugify(title)
    else:
        filename = str(filename)

    is_csv = str(filename).lower().endswith(".csv")
    if not is_csv and not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"

    # Reuse file_tools' sandbox: same workspace root, same traversal
    # rejection, so the deliverable lands next to the agent's other files.
    safe_path = _resolve_safe(filename)
    if safe_path is None:
        return {
            "status": "error",
            "output": f"Path rejected (must stay inside the workspace): {filename!r}",
        }

    normalized_rows, adjusted = _normalize_rows(headers, rows)
    if adjusted:
        print(
            f"[XLSX] WARNING: {adjusted} row(s) had a different length than "
            f"the headers ({len(headers)}); padded/truncated to fit."
        )

    try:
        safe_path.parent.mkdir(parents=True, exist_ok=True)
        if is_csv:
            import csv
            with open(safe_path, "w", newline="", encoding="utf-8") as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(headers)
                for r in normalized_rows:
                    writer.writerow(r)
            return {
                "status": "success",
                "output": f"CSV file saved as {filename}",
                "file_path": str(safe_path),
            }

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = title.strip()[:MAX_SHEET_TITLE]

        # Bold header row in row 1.
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        # One row per data entry.
        for row in normalized_rows:
            sheet.append(row)

        # Rough column-width heuristic: width ~ longest cell in the column,
        # capped so one giant value can't blow up the sheet.
        for column_cells in sheet.columns:
            longest = max(
                (len(str(cell.value)) for cell in column_cells if cell.value is not None),
                default=0,
            )
            sheet.column_dimensions[column_cells[0].column_letter].width = min(
                longest + 2, 50
            )

        # Append Provenance worksheet tab if requested or available
        try:
            from provenance.provenance_tracker import get_current_provenance_record
            if provenance_record is None and get_current_provenance_record is not None:
                provenance_record = get_current_provenance_record()

            if provenance_record and not is_csv:
                prov_sheet = workbook.create_sheet(title="Provenance")
                prov_sheet.append(["Field", "Audit Trail Detail"])
                for cell in prov_sheet[1]:
                    cell.font = Font(bold=True)

                sources = provenance_record.get("source_documents", [])
                sources_val = (
                    ", ".join(sources)
                    if sources
                    else "No external source documents used — general model knowledge"
                )

                prov_rows = [
                    ["Generated By", provenance_record.get("generated_by", "")],
                    ["Task Type", provenance_record.get("task_type", "")],
                    ["Timestamp", provenance_record.get("timestamp", "")],
                    ["Source Document(s)", sources_val],
                    ["Network Status", provenance_record.get("network_status", "")],
                    ["System", provenance_record.get("system", "")],
                ]
                for prow in prov_rows:
                    prov_sheet.append(prow)

                for column_cells in prov_sheet.columns:
                    longest = max(
                        (len(str(cell.value)) for cell in column_cells if cell.value is not None),
                        default=0,
                    )
                    prov_sheet.column_dimensions[column_cells[0].column_letter].width = min(
                        longest + 3, 70
                    )
        except Exception as prov_exc:
            print(f"[XLSX] Provenance sheet skipped: {prov_exc}")

        workbook.save(str(safe_path))
    except OSError as exc:
        return {"status": "error", "output": f"Could not save spreadsheet: {exc}"}
    except Exception as exc:  # noqa: BLE001 - library errors must not crash.
        return {"status": "error", "output": f"xlsx_generate failed: {exc}"}

    return {
        "status": "success",
        "output": f"Spreadsheet saved as {filename}",
        "file_path": str(safe_path),
    }


if __name__ == "__main__":
    TITLE = "Inspection Log"
    HEADERS = ["Item", "Finding", "Severity"]
    ROWS = [
        ["Joint A-12", "Minor corrosion", "Low"],
        ["Valve B-7", "No issues", "None"],
        ["Pump C-3", "Worn seal", "Medium"],
    ]

    print(f"Workspace: {WORKSPACE_DIR}\n")

    result = xlsx_generate(TITLE, HEADERS, ROWS)
    print("RESULT:", result)

    if result["status"] == "success":
        path = Path(result["file_path"])
        print(f"\nFile exists: {path.exists()} | Size: {path.stat().st_size} bytes")

        # Read the file back with openpyxl to confirm it is a real,
        # openable workbook with the right structure.
        from openpyxl import load_workbook

        wb = load_workbook(str(path))
        sheet = wb.active
        print(f"Sheet title: {sheet.title!r}")
        print(f"Dimensions: {sheet.dimensions}")
        for row in sheet.iter_rows(values_only=True):
            print(f"  {list(row)}")

    # Also exercise the padding behavior with a mismatched row.
    print("\nMISMATCHED ROW TEST (padding):")
    pad_result = xlsx_generate(
        "Pad Test",
        ["A", "B", "C"],
        [["1", "2", "3"], ["short"]],
    )
    print(f"  -> {pad_result}")
